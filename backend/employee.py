"""
employee.py — Employee Portal API.

Reuses the same database, models and AI pipeline output as the Customer
Portal; this router only adds aggregation/synthesis on top of data that
ai_agents.py already wrote (Feedback, AIInsight, MemoryRecord, RuntimeLog).
No new AI logic, no new database.
"""
import io
from collections import defaultdict
from datetime import datetime, timedelta

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy import func
from sqlalchemy.orm import Session

import models
import schemas
import auth
from database import get_db

router = APIRouter(prefix="/employee", tags=["employee"])
require_staff = auth.require_roles("employee")


# --- Dashboard ---------------------------------------------------------------

@router.get("/dashboard")
def dashboard(
    db: Session = Depends(get_db), user: models.User = Depends(require_staff),
    category: str | None = None,
):
    query = db.query(models.Feedback)
    if category:
        query = query.filter(models.Feedback.category == category)
    feedbacks = query.all()
    total = len(feedbacks)
    today = datetime.utcnow().date()
    feedback_today = sum(1 for f in feedbacks if f.created_at and f.created_at.date() == today)

    sentiment_counts = {"Positive": 0, "Negative": 0, "Neutral": 0}
    for f in feedbacks:
        if f.sentiment in sentiment_counts:
            sentiment_counts[f.sentiment] += 1

    positive_pct = round(sentiment_counts["Positive"] / total * 100, 1) if total else 0
    neutral_pct = round(sentiment_counts["Neutral"] / total * 100, 1) if total else 0
    negative_pct = round(sentiment_counts["Negative"] / total * 100, 1) if total else 0
    avg_sentiment = round(positive_pct - negative_pct, 1)  # simple net-sentiment score

    insight_query = db.query(models.AIInsight)
    memory_query = db.query(models.MemoryRecord)
    if category:
        insight_query = insight_query.filter(models.AIInsight.category == category)
        memory_query = memory_query.filter(models.MemoryRecord.category == category)

    products_count = db.query(func.count(models.Product.id)).scalar() or 0
    active_issues = insight_query.filter(models.AIInsight.status != "Implemented").count()
    recurring = memory_query.filter(models.MemoryRecord.occurrences > 1).count()
    recommendations = insight_query.count()

    runtime_logs = db.query(models.RuntimeLog).all()
    cost_saved = sum(l.cost_saved for l in runtime_logs)
    avg_latency_ms = 180  # simulated: deterministic local routing has no real network latency to measure

    # last-7-day sparkline of feedback volume
    sparkline = []
    for i in range(6, -1, -1):
        day = today - timedelta(days=i)
        sparkline.append(sum(1 for f in feedbacks if f.created_at and f.created_at.date() == day))

    return {
        "total_feedback": total,
        "feedback_today": feedback_today,
        "products": products_count,
        "average_sentiment": avg_sentiment,
        "positive_pct": positive_pct,
        "neutral_pct": neutral_pct,
        "negative_pct": negative_pct,
        "active_issues": active_issues,
        "recurring_complaints": recurring,
        "ai_recommendations": recommendations,
        "processing_time_ms": avg_latency_ms,
        "runtime_cost_saved": round(cost_saved, 4),
        "feedback_sparkline": sparkline,
    }


# --- Product Intelligence ----------------------------------------------------

def _product_health(db: Session, product: models.Product):
    feedbacks = db.query(models.Feedback).filter(models.Feedback.product_id == product.id).all()
    total = len(feedbacks)
    sentiment_counts = {"Positive": 0, "Negative": 0, "Neutral": 0}
    for f in feedbacks:
        if f.sentiment in sentiment_counts:
            sentiment_counts[f.sentiment] += 1

    satisfaction = round((sentiment_counts["Positive"] + 0.5 * sentiment_counts["Neutral"]) / total * 100) if total else 0
    pos_neg_ratio = round(sentiment_counts["Positive"] / max(sentiment_counts["Negative"], 1), 2)

    insights = (
        db.query(models.AIInsight)
        .filter(models.AIInsight.product_id == product.id)
        .order_by(models.AIInsight.frequency.desc())
        .all()
    )
    top_complaint = insights[0].problem if insights else None
    top_appreciated = "Fast, reliable experience" if sentiment_counts["Positive"] >= sentiment_counts["Negative"] else None

    # health score: blends satisfaction with complaint pressure
    complaint_pressure = min(40, sum(i.frequency for i in insights) * 4)
    health_score = max(0, min(100, satisfaction - complaint_pressure + 20))

    complaint_growth = "Rising" if insights and insights[0].frequency >= 3 else "Stable"

    return {
        "id": product.id,
        "name": product.name,
        "category": product.category,
        "health_score": health_score,
        "satisfaction": satisfaction,
        "total_feedback": total,
        "sentiment_counts": sentiment_counts,
        "pos_neg_ratio": pos_neg_ratio,
        "top_complaint": top_complaint,
        "top_appreciated_feature": top_appreciated,
        "complaint_growth": complaint_growth,
        "recommendation_status": insights[0].status if insights else "Generated",
    }


@router.get("/products")
def product_intelligence(db: Session = Depends(get_db), user: models.User = Depends(require_staff)):
    products = db.query(models.Product).all()
    return [_product_health(db, p) for p in products]


@router.get("/products/{product_id}")
def product_detail(product_id: int, db: Session = Depends(get_db), user: models.User = Depends(require_staff)):
    product = db.query(models.Product).filter(models.Product.id == product_id).first()
    if not product:
        raise HTTPException(404, "Product not found")

    health = _product_health(db, product)
    feedbacks = (
        db.query(models.Feedback)
        .filter(models.Feedback.product_id == product_id)
        .order_by(models.Feedback.created_at)
        .all()
    )

    sentiment_timeline = [
        {"date": f.created_at.strftime("%Y-%m-%d"), "sentiment": f.sentiment} for f in feedbacks
    ]

    category_counts = defaultdict(int)
    for f in feedbacks:
        if f.category:
            category_counts[f.category] += 1

    insights = (
        db.query(models.AIInsight)
        .filter(models.AIInsight.product_id == product_id)
        .order_by(models.AIInsight.frequency.desc())
        .all()
    )

    return {
        **health,
        "executive_summary": (
            f"{product.name} has received {health['total_feedback']} feedback entries with a "
            f"{health['satisfaction']}% satisfaction score. "
            + (f"The leading recurring issue is '{health['top_complaint']}'. " if health["top_complaint"] else "No major recurring issues detected. ")
            + ("Complaint volume is trending upward and should be prioritized." if health["complaint_growth"] == "Rising" else "Complaint volume is stable.")
        ),
        "sentiment_timeline": sentiment_timeline,
        "complaint_categories": dict(category_counts),
        "recurring_problems": [
            {"problem": i.problem, "frequency": i.frequency, "status": i.status} for i in insights
        ],
        "recommendation_timeline": [
            {"problem": i.problem, "recommendation": i.recommendation, "status": i.status, "updated_at": i.updated_at}
            for i in insights
        ],
    }


# --- Feedback Explorer ---------------------------------------------------------

@router.get("/feedback")
def feedback_explorer(
    db: Session = Depends(get_db),
    user: models.User = Depends(require_staff),
    product_id: int | None = None,
    language: str | None = None,
    sentiment: str | None = None,
    category: str | None = None,
    priority: str | None = None,
    q: str | None = None,
):
    query = db.query(models.Feedback)
    if product_id:
        query = query.filter(models.Feedback.product_id == product_id)
    if language:
        query = query.filter(models.Feedback.language == language)
    if sentiment:
        query = query.filter(models.Feedback.sentiment == sentiment)
    if category:
        query = query.filter(models.Feedback.category == category)
    if priority:
        query = query.filter(models.Feedback.priority == priority)
    if q:
        like = f"%{q}%"
        query = query.filter(
            (models.Feedback.original_text.ilike(like)) | (models.Feedback.translated_text.ilike(like))
        )

    rows = query.order_by(models.Feedback.created_at.desc()).all()
    products = {p.id: p.name for p in db.query(models.Product).all()}

    return [
        {
            "id": f.id,
            "product": products.get(f.product_id, "Unknown"),
            "product_id": f.product_id,
            "product_name": f.product_name,
            "image_path": f.image_path,
            "language": f.language,
            "original_text": f.original_text,
            "translated_text": f.translated_text,
            "sentiment": f.sentiment,
            "category": f.category,
            "problem": f.problem,
            "priority": f.priority,
            "recommendation": f.recommendation,
            "memory_note": f.memory_note,
            "created_at": f.created_at,
        }
        for f in rows
    ]


# --- Feedback Synthesizer (clustering) -----------------------------------------

@router.get("/clusters")
def feedback_synthesizer(db: Session = Depends(get_db), user: models.User = Depends(require_staff)):
    feedbacks = db.query(models.Feedback).filter(models.Feedback.problem.isnot(None)).all()
    products = {p.id: p.name for p in db.query(models.Product).all()}

    clusters = defaultdict(lambda: {"feedback": [], "products": set(), "languages": set()})
    for f in feedbacks:
        c = clusters[f.problem]
        c["feedback"].append(f)
        c["products"].add(products.get(f.product_id, "Unknown"))
        c["languages"].add(f.language)

    result = []
    for problem, data in clusters.items():
        count = len(data["feedback"])
        priorities = [f.priority for f in data["feedback"] if f.priority]
        priority = max(set(priorities), key=priorities.count) if priorities else "Low"
        confidence = min(98, 60 + count * 6)
        trend = "Rising" if count >= 3 else "Stable"
        recs = [f.recommendation for f in data["feedback"] if f.recommendation]

        result.append({
            "problem": problem,
            "related_feedback_count": count,
            "products": sorted(data["products"]),
            "languages": sorted(data["languages"]),
            "trend": trend,
            "confidence": confidence,
            "priority": priority,
            "business_impact": "High" if count >= 4 else "Medium" if count >= 2 else "Low",
            "recommendation": recs[0] if recs else None,
        })

    result.sort(key=lambda c: c["related_feedback_count"], reverse=True)
    return result


# --- AI Recommendation Center ---------------------------------------------------

@router.get("/recommendations")
def recommendation_center(db: Session = Depends(get_db), user: models.User = Depends(require_staff)):
    insights = db.query(models.AIInsight).order_by(models.AIInsight.frequency.desc()).all()
    products = {p.id: p.name for p in db.query(models.Product).all()}

    return [
        {
            "id": i.id,
            "recommendation": i.recommendation,
            "product": products.get(i.product_id, "Unknown"),
            "product_id": i.product_id,
            "confidence": min(98, 55 + i.frequency * 7),
            "supporting_feedback": i.frequency,
            "business_impact": "High" if i.impact_score >= 20 else "Medium" if i.impact_score >= 10 else "Low",
            "estimated_customer_impact": i.frequency * 12,  # rough proxy: affected-customer estimate
            "priority": "High" if i.impact_score >= 20 else "Medium" if i.impact_score >= 10 else "Low",
            "status": i.status,
            "updated_at": i.updated_at,
        }
        for i in insights
    ]


@router.patch("/recommendations/{insight_id}/status")
def update_recommendation_status(
    insight_id: int, payload: schemas.RecommendationStatusUpdate,
    db: Session = Depends(get_db), user: models.User = Depends(require_staff),
):
    if payload.status not in ("Generated", "Accepted", "Rejected", "Implemented"):
        raise HTTPException(400, "Invalid status")
    insight = db.query(models.AIInsight).filter(models.AIInsight.id == insight_id).first()
    if not insight:
        raise HTTPException(404, "Recommendation not found")
    insight.status = payload.status
    insight.updated_at = datetime.utcnow()
    db.commit()
    return {"id": insight.id, "status": insight.status}


# --- Memory Intelligence ------------------------------------------------------

@router.get("/memory")
def memory_intelligence(db: Session = Depends(get_db), user: models.User = Depends(require_staff)):
    records = db.query(models.MemoryRecord).order_by(models.MemoryRecord.occurrences.desc()).all()
    products = {p.id: p.name for p in db.query(models.Product).all()}

    return [
        {
            "id": r.id,
            "product": products.get(r.product_id, "Unknown"),
            "problem": r.problem_key,
            "occurrences": r.occurrences,
            "similarity_score": min(99, 70 + r.occurrences * 4),
            "frequency_growth": f"+{r.occurrences - 1}" if r.occurrences > 1 else "New",
            "first_seen": r.first_seen,
            "latest_seen": r.updated_at,
            "ai_conclusion": (
                f"This issue has recurred {r.occurrences} times for {products.get(r.product_id, 'this product')} "
                "and should be treated as a systemic problem rather than an isolated complaint."
                if r.occurrences > 2
                else r.trend_note
            ),
        }
        for r in records
    ]


# --- Trend Analytics ------------------------------------------------------------

@router.get("/trends")
def trend_analytics(
    db: Session = Depends(get_db), user: models.User = Depends(require_staff),
    category: str | None = None,
):
    query = db.query(models.Feedback)
    if category:
        query = query.filter(models.Feedback.category == category)
    feedbacks = query.all()
    products = {p.id: p.name for p in db.query(models.Product).all()}

    sentiment_counts = {"Positive": 0, "Negative": 0, "Neutral": 0}
    category_counts = defaultdict(int)
    language_counts = defaultdict(int)
    product_counts = defaultdict(int)
    volume_by_day = defaultdict(int)

    for f in feedbacks:
        if f.sentiment in sentiment_counts:
            sentiment_counts[f.sentiment] += 1
        if f.category:
            category_counts[f.category] += 1
        language_counts[f.language] += 1
        product_counts[products.get(f.product_id, "Unknown")] += 1
        if f.created_at:
            volume_by_day[f.created_at.strftime("%Y-%m-%d")] += 1

    weekly = sorted(volume_by_day.items())[-7:]
    total = len(feedbacks)
    satisfaction_trend = []
    running_pos, running_total = 0, 0
    for day, count in weekly:
        day_feedbacks = [f for f in feedbacks if f.created_at and f.created_at.strftime("%Y-%m-%d") == day]
        running_pos += sum(1 for f in day_feedbacks if f.sentiment == "Positive")
        running_total += len(day_feedbacks)
        satisfaction_trend.append({
            "date": day,
            "satisfaction": round(running_pos / running_total * 100) if running_total else 0,
        })

    return {
        "sentiment": sentiment_counts,
        "feedback_volume_by_day": dict(weekly),
        "product_comparison": dict(product_counts),
        "language_distribution": dict(language_counts),
        "complaint_categories": dict(category_counts),
        "satisfaction_trend": satisfaction_trend,
        "total_feedback": total,
    }


# --- Executive Summary -----------------------------------------------------------

@router.get("/executive-summary")
def executive_summary(db: Session = Depends(get_db), user: models.User = Depends(require_staff)):
    feedbacks = db.query(models.Feedback).all()
    total = len(feedbacks)
    sentiment_counts = {"Positive": 0, "Negative": 0, "Neutral": 0}
    for f in feedbacks:
        if f.sentiment in sentiment_counts:
            sentiment_counts[f.sentiment] += 1
    satisfaction = round((sentiment_counts["Positive"] + 0.5 * sentiment_counts["Neutral"]) / total * 100) if total else 0

    insights = db.query(models.AIInsight).order_by(models.AIInsight.frequency.desc()).limit(3).all()
    products = {p.id: p.name for p in db.query(models.Product).all()}
    attention_products = sorted(
        {products.get(i.product_id, "Unknown") for i in insights}
    )

    headline = (
        f"Customer satisfaction stands at {satisfaction}% across {total} processed feedback items. "
        + (f"{len(insights)} recurring issue(s) require attention, most notably "
           + ", ".join(f"'{i.problem}'" for i in insights) + "." if insights else "No major recurring issues currently flagged.")
    )

    return {
        "generated_at": datetime.utcnow(),
        "headline": headline,
        "satisfaction": satisfaction,
        "total_feedback": total,
        "sentiment_counts": sentiment_counts,
        "top_recurring_issues": [
            {"problem": i.problem, "frequency": i.frequency, "recommendation": i.recommendation} for i in insights
        ],
        "products_requiring_attention": attention_products,
        "recommendation_count": db.query(func.count(models.AIInsight.id)).scalar() or 0,
    }


# --- Notifications ---------------------------------------------------------------

@router.get("/notifications")
def notifications(db: Session = Depends(get_db), user: models.User = Depends(require_staff)):
    items = []
    insights = db.query(models.AIInsight).all()
    for i in insights:
        if i.frequency >= 4:
            items.append({
                "id": f"spike-{i.id}",
                "type": "complaint_spike",
                "title": "Complaint spike detected",
                "message": f"'{i.problem}' has been reported {i.frequency} times.",
                "severity": "high",
                "created_at": i.updated_at,
            })

    products = db.query(models.Product).all()
    for p in products:
        health = _product_health(db, p)
        if health["health_score"] < 50 and health["total_feedback"] > 0:
            items.append({
                "id": f"health-{p.id}",
                "type": "product_health_decline",
                "title": f"{p.name} health declining",
                "message": f"{p.name}'s health score has dropped to {health['health_score']}.",
                "severity": "medium",
                "created_at": datetime.utcnow(),
            })

    recent_insights = db.query(models.AIInsight).order_by(models.AIInsight.updated_at.desc()).limit(5).all()
    for i in recent_insights:
        items.append({
            "id": f"rec-{i.id}",
            "type": "recommendation_generated",
            "title": "New AI recommendation",
            "message": f"Recommendation generated for '{i.problem}'.",
            "severity": "low",
            "created_at": i.updated_at,
        })

    memory_growth = db.query(models.MemoryRecord).filter(models.MemoryRecord.occurrences >= 3).all()
    for m in memory_growth:
        items.append({
            "id": f"memory-{m.id}",
            "type": "memory_cluster_growth",
            "title": "Memory cluster growing",
            "message": f"'{m.problem_key}' memory cluster has grown to {m.occurrences} occurrences.",
            "severity": "medium",
            "created_at": m.updated_at,
        })

    items.sort(key=lambda n: n["created_at"], reverse=True)
    return items[:30]


# --- Category Insights -------------------------------------------------------------

def _category_aggregates(db: Session):
    """Shared helper: aggregates Feedback + AIInsight per customer-selected category."""
    feedbacks = db.query(models.Feedback).filter(models.Feedback.category.isnot(None)).all()
    insights = db.query(models.AIInsight).all()

    by_category = defaultdict(lambda: {"feedback": [], "insights": []})
    for f in feedbacks:
        by_category[f.category]["feedback"].append(f)
    for i in insights:
        if i.category:
            by_category[i.category]["insights"].append(i)

    results = {}
    for category, data in by_category.items():
        fbs = data["feedback"]
        total = len(fbs)
        sentiment_counts = {"Positive": 0, "Negative": 0, "Neutral": 0}
        for f in fbs:
            if f.sentiment in sentiment_counts:
                sentiment_counts[f.sentiment] += 1
        positive_pct = round(sentiment_counts["Positive"] / total * 100, 1) if total else 0
        negative_pct = round(sentiment_counts["Negative"] / total * 100, 1) if total else 0

        insights_sorted = sorted(data["insights"], key=lambda i: i.frequency, reverse=True)
        results[category] = {
            "category": category,
            "total_feedback": total,
            "sentiment_counts": sentiment_counts,
            "positive_pct": positive_pct,
            "negative_pct": negative_pct,
            "net_sentiment": round(positive_pct - negative_pct, 1),
            "insights": insights_sorted,
        }
    return results


@router.get("/categories")
def category_insights(db: Session = Depends(get_db), user: models.User = Depends(require_staff)):
    aggregates = _category_aggregates(db)
    out = []
    for category, data in aggregates.items():
        top_insight = data["insights"][0] if data["insights"] else None
        out.append({
            "category": category,
            "total_feedback": data["total_feedback"],
            "sentiment_counts": data["sentiment_counts"],
            "positive_pct": data["positive_pct"],
            "negative_pct": data["negative_pct"],
            "recurring_issues": [
                {"problem": i.problem, "frequency": i.frequency, "status": i.status} for i in data["insights"]
            ],
            "top_recommendation": top_insight.recommendation if top_insight else None,
        })
    out.sort(key=lambda c: c["total_feedback"], reverse=True)
    return out


# --- Company Improvement Intelligence -----------------------------------------------

def _priority_from_impact(score: float) -> str:
    return "High" if score >= 20 else "Medium" if score >= 10 else "Low"


@router.get("/improvement-intelligence")
def improvement_intelligence(db: Session = Depends(get_db), user: models.User = Depends(require_staff)):
    aggregates = _category_aggregates(db)

    strengths, improvements = [], []
    for category, data in aggregates.items():
        if data["total_feedback"] == 0:
            continue

        if data["net_sentiment"] > 0:
            strengths.append({
                "category": category,
                "positive_pct": data["positive_pct"],
                "total_feedback": data["total_feedback"],
                "highlight": (
                    f"{data['positive_pct']}% positive feedback — customers consistently appreciate "
                    f"the {category.lower()} experience."
                ),
            })
        elif data["net_sentiment"] < 0 and data["insights"]:
            worst = data["insights"][0]  # highest-frequency recurring problem in this category
            improvements.append({
                "category": category,
                "root_cause": worst.problem,
                "supporting_feedback_count": worst.frequency,
                "business_impact": _priority_from_impact(worst.impact_score),
                "priority": _priority_from_impact(worst.impact_score),
                "ai_confidence": min(98, 55 + worst.frequency * 7),
                "recommended_action": worst.recommendation,
                "negative_pct": data["negative_pct"],
                "total_feedback": data["total_feedback"],
            })

    strengths.sort(key=lambda s: s["positive_pct"], reverse=True)
    improvements.sort(key=lambda i: i["supporting_feedback_count"], reverse=True)

    return {
        "strengths": strengths[:6],
        "improvements": improvements[:6],
    }


# --- Reports (Excel) --------------------------------------------------------------

HEADER_FILL = PatternFill(start_color="4B4A8E", end_color="4B4A8E", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _write_sheet(ws, headers, rows):
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append(row)
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(60, max(12, length + 2))
    ws.freeze_panes = "A2"


@router.get("/reports/excel")
def export_excel(
    report_type: str = Query("feedback", description="feedback | recommendations | runtime | routing_breakdown"),
    db: Session = Depends(get_db), user: models.User = Depends(require_staff),
):
    products = {p.id: p.name for p in db.query(models.Product).all()}
    wb = Workbook()
    ws = wb.active

    if report_type == "routing_breakdown":
        # Mirrors exactly what the Runtime Monitor pie chart shows: the same
        # most-recent-100 logs, grouped by task, plus the full per-log detail.
        logs = db.query(models.RuntimeLog).order_by(models.RuntimeLog.created_at.desc()).limit(100).all()
        total = len(logs) or 1

        by_task = defaultdict(list)
        for l in logs:
            by_task[l.task].append(l)

        ws.title = "Pie Chart Breakdown"
        summary_rows = []
        for task, task_logs in sorted(by_task.items(), key=lambda kv: len(kv[1]), reverse=True):
            count = len(task_logs)
            total_cost = sum(l.cost for l in task_logs)
            total_saved = sum(l.cost_saved for l in task_logs)
            models_used = ", ".join(sorted({l.model for l in task_logs}))
            summary_rows.append([
                task, count, round(count / total * 100, 1),
                round(total_cost, 4), round(total_saved, 4),
                round(total_cost / count, 4) if count else 0, models_used,
            ])
        _write_sheet(
            ws,
            ["Task", "Count", "Percentage", "Total Cost", "Total Cost Saved", "Avg Cost", "Models Used"],
            summary_rows,
        )

        ws2 = wb.create_sheet("Log Detail")
        detail_rows = [
            [l.task, l.model, l.reason, l.cost, l.cost_saved, l.created_at.strftime("%Y-%m-%d %H:%M:%S") if l.created_at else ""]
            for l in logs
        ]
        _write_sheet(ws2, ["Task", "Model", "Reason", "Cost", "Cost Saved", "Created At"], detail_rows)

    elif report_type == "recommendations":
        ws.title = "Recommendations"
        rows = [
            [products.get(i.product_id, ""), i.problem, i.frequency, i.impact_score, i.status, i.recommendation]
            for i in db.query(models.AIInsight).all()
        ]
        _write_sheet(ws, ["Product", "Problem", "Frequency", "Impact Score", "Status", "Recommendation"], rows)
    elif report_type == "runtime":
        ws.title = "Runtime"
        rows = [
            [l.task, l.model, l.reason, l.cost, l.cost_saved, l.created_at.strftime("%Y-%m-%d %H:%M:%S") if l.created_at else ""]
            for l in db.query(models.RuntimeLog).all()
        ]
        _write_sheet(ws, ["Task", "Model", "Reason", "Cost", "Cost Saved", "Created At"], rows)
    else:
        ws.title = "Feedback"
        rows = [
            [
                products.get(f.product_id, ""), f.language, f.original_text,
                f.sentiment, f.category, f.priority, f.recommendation,
                f.created_at.strftime("%Y-%m-%d %H:%M:%S") if f.created_at else "",
            ]
            for f in db.query(models.Feedback).all()
        ]
        _write_sheet(ws, ["Product", "Language", "Original Text", "Sentiment", "Category", "Priority", "Recommendation", "Created At"], rows)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={report_type}_report.xlsx"},
    )
